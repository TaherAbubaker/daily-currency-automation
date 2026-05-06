"""
Currency Rate Agent
===================
Fetches USD → JOD and USD → ILS (NIS) rates daily and appends to Excel.
Run once manually or schedule with cron / Task Scheduler.
"""

import os
import requests
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_FILE = "currency_rates.xlsx"
API_URL = "https://cdn.jsdelivr.net/npm/@fawazahmed0/currency-api@latest/v1/currencies/usd.json"
FALLBACK_URL = "https://latest.currency-api.pages.dev/v1/currencies/usd.json"


def fetch_rates():
    """Fetch USD to JOD and ILS rates from free API (no key needed)."""
    for url in [API_URL, FALLBACK_URL]:
        try:
            r = requests.get(url, timeout=10)
            if r.status_code == 200:
                data = r.json()
                usd_rates = data.get("usd", {})
                jod = usd_rates.get("jod")
                ils = usd_rates.get("ils")  # ILS = NIS (Israeli New Shekel)
                if jod and ils:
                    return round(jod, 4), round(ils, 4)
        except Exception as e:
            print(f"  ⚠ Tried {url}: {e}")
    return None, None


def setup_workbook():
    """Create a new Excel file with headers and formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Currency Rates"

    # --- Header row ---
    headers = ["Date", "Day", "USD → JOD", "USD → NIS (ILS)", "Note"]
    ws.append(headers)

    # Style the header
    header_fill = PatternFill("solid", start_color="1F4E79")  # Dark blue
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Column widths
    widths = [14, 12, 14, 18, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"  # Freeze header row

    wb.save(EXCEL_FILE)
    print(f"  ✅ Created new file: {EXCEL_FILE}")
    return wb


def append_rate_row(jod_rate, ils_rate, note=""):
    """Append today's rates as a new row."""
    if not os.path.exists(EXCEL_FILE):
        setup_workbook()

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    day_name = now.strftime("%A")

    # Check if today's row already exists
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == date_str:
            print(f"  ℹ  Rate for {date_str} already logged. Skipping.")
            wb.close()
            return

    # Row styling alternates
    row_num = ws.max_row + 1
    is_even = (row_num % 2 == 0)
    row_fill = PatternFill("solid", start_color="D6E4F0" if is_even else "FFFFFF")

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    new_row = [date_str, day_name, jod_rate, ils_rate, note]

    for col_idx, value in enumerate(new_row, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        if col_idx == 1:
            cell.font = Font(name="Arial", bold=True, size=10)
        else:
            cell.font = Font(name="Arial", size=10)

    # Add auto-filter if not already on
    if not ws.auto_filter.ref:
        ws.auto_filter.ref = f"A1:E1"

    wb.save(EXCEL_FILE)
    print(f"  ✅ Logged: {date_str} | 1 USD = {jod_rate} JOD | 1 USD = {ils_rate} NIS")


def run_agent():
    print(f"\n🤖 Currency Agent Running — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("  📡 Fetching exchange rates...")

    jod, ils = fetch_rates()

    if jod and ils:
        append_rate_row(jod, ils)
    else:
        print("  ❌ Could not fetch rates. Check your internet connection.")
        # Log failure row so we know it ran
        append_rate_row("N/A", "N/A", note="Fetch failed - no internet?")


if __name__ == "__main__":
    run_agent()